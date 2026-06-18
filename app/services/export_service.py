from io import BytesIO
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime

class ExportService:
    @staticmethod
    def export_to_excel(data: list, columns: list, filename: str = "report.xlsx"):
        """Export data to Excel using openpyxl directly"""
        if not data:
            return None
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Add headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_idx, row in enumerate(data, 2):
                for col_idx, column in enumerate(columns, 1):
                    value = row.get(column, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(columns, 1):
                max_length = max(
                    len(str(column)),
                    max([len(str(row.get(column, ''))) for row in data[:100]]) if data else 0
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            # Fallback to CSV if openpyxl fails
            return ExportService._export_to_csv(data, columns)
    
    @staticmethod
    def _export_to_csv(data: list, columns: list):
        """Fallback export to CSV"""
        output = BytesIO()
        # Write BOM for UTF-8
        output.write(b'\xef\xbb\xbf')
        # Write CSV
        csv_output = BytesIO()
        import csv
        writer = csv.writer(csv_output)
        writer.writerow(columns)
        for row in data:
            writer.writerow([row.get(col, '') for col in columns])
        output.write(csv_output.getvalue())
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_pdf(data: list, columns: list, title: str = "Report"):
        """Export data to PDF"""
        if not data:
            return None
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#f97316'),
            alignment=1,  # Center
            spaceAfter=30
        )
        
        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.25*inch))
        
        # Add date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=1
        )
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        elements.append(Spacer(1, 0.25*inch))
        
        # Table data
        table_data = [columns]
        for row in data[:1000]:  # Limit to 1000 rows for PDF
            table_data.append([str(row.get(col, '')) for col in columns])
        
        # Table style
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output